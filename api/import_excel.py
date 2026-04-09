"""api/import_excel.py — импорт заказов из xlsx/csv через Integram (fallback: openpyxl)."""

import json
import logging
import time
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile

from integram.client import IntegramClient
from integram.deps import get_integram
from integram.exceptions import IntegramError
from services.client_service import find_or_create

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["import"])

_MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
_XLSX_MAGIC = b"PK\x03\x04"  # ZIP/XLSX magic bytes

# Rate limiting: не более 10 запросов в минуту на IP
_RATE_LIMIT = 10
_RATE_WINDOW = 60.0
_rate_store: dict[str, list[float]] = defaultdict(list)


async def _check_rate_limit(request: Request) -> None:
    ip = request.client.host if request.client else "unknown"
    now = time.monotonic()
    cutoff = now - _RATE_WINDOW
    timestamps = [t for t in _rate_store[ip] if t > cutoff]
    if len(timestamps) >= _RATE_LIMIT:
        raise HTTPException(429, "Слишком много запросов. Повторите через минуту.")
    timestamps.append(now)
    _rate_store[ip] = timestamps


def _validate_file(file: UploadFile) -> None:
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".csv")):
        raise HTTPException(400, "Разрешены только .xlsx, .xls, .csv файлы")


def _validate_mime(content: bytes, filename: str) -> None:
    """Проверка magic bytes для xlsx/xls (ZIP-формат). CSV — пропускаем."""
    if filename.lower().endswith(".csv"):
        return
    if not content.startswith(_XLSX_MAGIC):
        raise HTTPException(400, "Файл не является валидным xlsx/xls (неверная сигнатура)")


# ── Preview ──────────────────────────────────────────────────────────────────


@router.post("/excel/preview")
async def import_preview(
    file: UploadFile,
    igm: IntegramClient = Depends(get_integram),
    _rl: None = Depends(_check_rate_limit),
):
    """Вернуть заголовки и первые строки файла для настройки маппинга."""
    _validate_file(file)
    content = await file.read()
    if len(content) > _MAX_FILE_BYTES:
        raise HTTPException(413, "Файл превышает 10 МБ")
    _validate_mime(content, file.filename or "")
    try:
        return await igm.import_preview(content, file.filename or "import.xlsx")
    except IntegramError as exc:
        logger.error("Integram preview failed: %s", exc)
        # Fallback: опарсить локально через openpyxl
        return _local_preview(content, file.filename or "")


def _local_preview(content: bytes, filename: str) -> dict:
    """Локальный предпросмотр через openpyxl (fallback)."""
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"headers": [], "preview": [], "totalRows": 0}
        headers = [str(c or "") for c in rows[0]]
        preview = [[str(c or "") for c in r] for r in rows[1:11]]
        return {"headers": headers, "preview": preview, "totalRows": len(rows) - 1}
    except Exception as exc:
        raise HTTPException(500, f"Не удалось прочитать файл: {exc}")


# ── Import ────────────────────────────────────────────────────────────────────


@router.post("/excel")
async def import_excel(
    file: UploadFile,
    mapping: str | None = None,
    igm: IntegramClient = Depends(get_integram),
    _rl: None = Depends(_check_rate_limit),
):
    """
    Импортировать заказы из xlsx/csv в Integram (typeId=17).

    mapping — JSON строка { "fileColIndex": "colId" | "__val__" | "__skip__" }
    Пример: {"0": "__val__", "1": "33", "2": "__skip__"}

    Колонка с именем/телефоном клиента автоматически обрабатывается через find_or_create.
    """
    _validate_file(file)
    content = await file.read()
    if len(content) > _MAX_FILE_BYTES:
        raise HTTPException(413, "Файл превышает 10 МБ")
    _validate_mime(content, file.filename or "")

    parsed_mapping: dict[str, str] | None = None
    if mapping:
        try:
            parsed_mapping = json.loads(mapping)
            _validate_mapping(parsed_mapping)
        except (json.JSONDecodeError, ValueError) as exc:
            raise HTTPException(400, f"Неверный mapping: {exc}")

    # Попытка через Integram
    try:
        result = await igm.import_xlsx(
            content,
            file.filename or "import.xlsx",
            typeId=IntegramClient.T_ORDERS,
            mapping=parsed_mapping,
        )
        return {"source": "integram", **result}
    except IntegramError as exc:
        logger.warning("Integram import failed, falling back to openpyxl: %s", exc)

    # Fallback через openpyxl
    return await _local_import(igm, content, file.filename or "")


def _validate_mapping(mapping: dict) -> None:
    """Разрешаем только colId из белого списка + служебные ключи."""
    allowed_col_ids = {
        "30", "31", "32", "33", "34", "35",  # колонки Заказов
        "__val__", "__skip__",
    }
    for val in mapping.values():
        if val not in allowed_col_ids:
            raise ValueError(f"Недопустимый colId в mapping: {val!r}")


async def _local_import(igm: IntegramClient, content: bytes, filename: str) -> dict:
    """Построчный fallback через openpyxl → find_or_create → create_order."""
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(500, f"Не удалось прочитать файл: {exc}")

    if len(rows) < 2:
        raise HTTPException(400, "Файл пустой или нет строк данных")

    headers = [str(c or "").strip().lower() for c in rows[0]]
    created, errors = [], []

    for i, row in enumerate(rows[1:], start=2):
        try:
            data = {headers[j]: str(v or "").strip() for j, v in enumerate(row)}

            # Дедупликация клиента
            phone = data.get("phone") or data.get("телефон") or data.get("тел")
            email = data.get("email") or data.get("e-mail")
            name = data.get("name") or data.get("имя") or data.get("клиент")

            if not (phone or email):
                errors.append({"row": i, "message": "Нет phone или email"})
                continue

            client, _ = await find_or_create(igm, phone=phone, email=email, name=name)

            # Создать заказ
            order = await igm.create_order_with_event(
                client_id=client["id"],
                status_id=IntegramClient.STATUS_MAP["NEW"],
                source="TABLE",
                payload=data,
            )
            created.append(order["id"])
        except Exception as exc:
            errors.append({"row": i, "message": str(exc)})

    return {
        "source": "openpyxl",
        "created": len(created),
        "ids": created,
        "errors": errors,
    }
